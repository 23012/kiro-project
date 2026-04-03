import openpyxl, json

wb = openpyxl.load_workbook(r'cau-drug-support/drug-list.xlsx')
ws = wb['개인 투약 이력']

rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    date_str = str(row[0])
    rows.append({
        "date": f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}",
        "code": str(row[1]),
        "name": row[2],
        "component": row[3],
        "dose": row[4],
        "unit": row[5],
        "frequency": row[6],
        "days": row[7],
        "componentCode": row[8]
    })

with open(r'cau-drug-support/js/drug-data.js', 'w', encoding='utf-8') as f:
    f.write('const DRUG_DATA = ')
    json.dump(rows, f, ensure_ascii=False, indent=2)
    f.write(';\n')

print(f'Converted {len(rows)} records')
