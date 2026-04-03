import openpyxl
wb = openpyxl.load_workbook(r'cau-drug-support/drug-list.xlsx')
ws = wb['개인 투약 이력']
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
    print([cell.value for cell in row])
